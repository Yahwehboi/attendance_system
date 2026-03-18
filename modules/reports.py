import os
import sys
from datetime import datetime

sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from database.db_setup import create_connection

# ── Path setup ────────────────────────────────────────────────────────────────
if getattr(sys, 'frozen', False):
    APP_DIR = os.path.dirname(sys.executable)
else:
    APP_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

REPORTS_DIR = os.path.join(APP_DIR, 'reports')
os.makedirs(REPORTS_DIR, exist_ok=True)


def export_to_excel(date_filter=None, course_id=None):
    """Export attendance to Excel, optionally filtered by date or course."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment

        conn = create_connection()
        cursor = conn.cursor()

        query = '''
            SELECT s.student_id, s.name, s.department, s.level,
                   a.date, a.time, a.status
            FROM attendance a
            JOIN students s ON a.student_id = s.student_id
            WHERE 1=1
        '''
        params = []
        if date_filter:
            query += " AND a.date = ?"
            params.append(date_filter)
        if course_id:
            query += " AND a.course_id = ?"
            params.append(course_id)
        query += " ORDER BY a.date DESC, a.time ASC"

        cursor.execute(query, params)
        records = cursor.fetchall()
        conn.close()

        if not records:
            return False, "❌ No attendance records found."

        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Attendance Report"

        # Header styling
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="7c3aed")
        headers = ["Student ID", "Name", "Department",
                   "Level", "Date", "Time", "Status"]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        # Data rows
        for row_idx, record in enumerate(records, 2):
            for col_idx, value in enumerate(record, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = Alignment(horizontal="center")
                if row_idx % 2 == 0:
                    cell.fill = PatternFill("solid", fgColor="F3F0FF")

        # Auto column width
        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = max_len + 4

        # Save file
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename  = os.path.join(REPORTS_DIR,
                                 f"attendance_{timestamp}.xlsx")
        wb.save(filename)
        return True, f"✅ Report saved:\n{filename}"

    except Exception as e:
        return False, f"❌ Error: {str(e)}"


def get_summary_report():
    """Get total attendance count per student."""
    conn = create_connection()
    cursor = conn.cursor()
    cursor.execute('''
        SELECT s.student_id, s.name, s.department, s.level,
               COUNT(a.id) as total_present
        FROM students s
        LEFT JOIN attendance a ON s.student_id = a.student_id
        GROUP BY s.student_id
        ORDER BY s.name ASC
    ''')
    records = cursor.fetchall()
    conn.close()
    return records


def get_monthly_report(year, month):
    """Get attendance records for a specific month."""
    conn = create_connection()
    cursor = conn.cursor()
    month_str = f"{year}-{str(month).zfill(2)}"
    cursor.execute('''
        SELECT s.student_id, s.name, s.department, s.level,
               a.date, a.time, a.status
        FROM attendance a
        JOIN students s ON a.student_id = s.student_id
        WHERE a.date LIKE ?
        ORDER BY a.date ASC, a.time ASC
    ''', (f"{month_str}%",))
    records = cursor.fetchall()
    conn.close()
    return records